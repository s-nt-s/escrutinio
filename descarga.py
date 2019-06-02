#!/usr/bin/env python3
import os
import zipfile
from io import BytesIO
from urllib.parse import urljoin, urlparse

import bs4
import requests
import yaml
import argparse
import sys

from bunch import Bunch
from openpyxl import load_workbook

parser = argparse.ArgumentParser(description='Descarga y analiza los ficheros del escrutinio')
parser.add_argument('--descarga', action='store_true', help='Descarga los ficheros')
parser.add_argument('--analiza', action='store_true', help='Analiza los ficheros')
parser.add_argument('--machaca', action='store_true', help='Machaca los ficheros ya creados')

args = parser.parse_args()

url = "http://www.infoelectoral.mir.es/infoelectoral/min/areaDescarga.html?method=inicio"

def load(file):
    with open(file, "r") as f:
        r = yaml.safe_load_all(f)
        return [Bunch(**i) for i in r]

def get(url):
    r = requests.get(url)
    soup = bs4.BeautifulSoup(r.content, "lxml")
    for a in soup.findAll("a"):
        href = a.attrs.get("href", "#")
        if not href.startswith("#"):
            href = urljoin(url, href)
            a.attrs["href"] = href
    return soup


def get_info(a):
    url = a.attrs["href"]
    purl = urlparse(url)
    file = os.path.basename(purl.path)
    geo = "muni"
    if file.startswith("PROV_"):
        file = file[5:]
        geo = "prov"
    tp = int(file[:2])
    if tp == 1:
        tp = "referendum"
    elif tp == 2:
        tp = "congreso"
    elif tp == 3:
        tp = "senado"
    elif tp == 4:
        tp = "municipal"
    elif tp == 6:
        tp = "cabildo"
    elif tp == 7:
        tp = "europeas"
    year = int(file[3:7])
    month = int(file[7:9])
    day = int(a.get_text().strip().split()[0])
    date = "%04d.%02d.%02d" % (year, month, day)
    return (tp, date, geo, url)


def get_zip(url):
    r = requests.get(url)
    bytes = BytesIO(r.content)
    with zipfile.ZipFile(bytes) as thezip:
        files = thezip.infolist()
        if len(files) == 0:
            raise Exception("No hay ficheros en %s" % (url))
        if len(files) > 1:
            raise Exception("Demasidados ficheros en %s\n%s" %
                            (url, ", ".join(z.filename for z in files)))
        file = files[0]
        with thezip.open(file) as thefile:
            return (thefile.read(), file.filename, os.path.splitext(file.filename)[1])

def trimRow(row):
    for i, r in enumerate(row):
        if r is not None:
            return (i, [r for r in row[i:] if r is not None])

def get_rows(sheet, ini=1, max_column=None, must=None):
    max_row=sheet.max_row
    if max_column is None:
        max_column=sheet.max_column
    i = ini - 1
    for cells in sheet.iter_rows(min_row=ini, max_col=max_column):
        i = i + 1
        row=[]
        notNull = 0
        frsVal = None
        for cell_obj in cells:
            val = cell_obj.value
            if isinstance(val, str):
                val = val.strip()
                if val.isdigit():
                    val = int(val)
            if val is not None:
                notNull = notNull + 1
                if frsVal is None:
                    frsVal = val
            row.append(val)
        if notNull>1 and frsVal!="Total":
            while len(row)>0 and row[-1] is None:
                row.pop()
            if must is None or (must<len(row) and row[must] is not None):
                yield i, row


def parse_head_xlsx(sheet):
    partidos=[]
    abreviat=[]
    for i, row in get_rows(sheet):
        if row[0] is not None:
            return i, partidos, abreviat, row
        if len(partidos)==0:
            partidos = trimRow(row)
            continue
        if len(abreviat)==0:
            abreviat = trimRow(row)
            continue

def parse_xlsx(archivo):
    print(archivo, end="\r")
    circunscripciones=[]
    wb = load_workbook(filename=archivo, read_only=True)
    sheet = wb.active
    irow, partidos, abreviat, head = parse_head_xlsx(sheet)
    ini, partidos = partidos
    fields={}
    for k, c in (
          ("codcir", "Código de Provincia"),
          ("codmun", "Código de Municipio"),
          ("comunidad", "Nombre de Comunidad"),
          ("provincia", "Nombre de Provincia"),
          ("municipio", "Nombre de Municipio"),
          ("blancos", "Votos en blanco"),
          ("censo", "Total censo electoral"),
          ("nulos", "Votos nulos"),
          ("validos", "Votos válidos"),
          ("votos", "Total votantes"),
    ):
        if c in head:
            fields[k]=head.index(c)
    isDiputados = "Diputados" in head
    diputados = 0
    fin = len(head)
    js = range(ini, fin, 2) if isDiputados else range(ini, fin)
    js = list(js)
    for i, row in get_rows(sheet, ini=irow+1, max_column=fin, must=0):
        print(archivo, i-irow, end="\r")
        iP = -1
        cir = {k:row[v] for k,v in fields.items()}
        cir["abstencion"]=cir["censo"]-cir["votos"]
        cir["partidos"]={}
        for j in js:
            iP = iP + 1
            p = partidos[iP]
            votos = row[j]
            if votos>0:
                if isDiputados:
                    cir["diputados"] = cir.get("diputados", 0) + row[j+1]
                cir["partidos"][p]=votos
        circunscripciones.append(cir)
    print(archivo)
    return sorted(circunscripciones, key=lambda x: x["codcir"])

if args.descarga:
    data = {}
    soup = get(url)
    for i, a in enumerate(soup.select("table.data.candidatos a")):
        tp, date, geo, url = get_info(a)
        dr = "data/%s/" % tp
        os.makedirs(dr, exist_ok=True)
        file = "%s_%s" % (geo, date)
        content, name, ext = get_zip(url)
        out = dr+file+ext
        with open(out, "wb") as f:
            f.write(content)
        key = (tp, date)
        items = data.get(key, [])
        items.append((geo, url, out))
        data[key] = items

    with open("index.yml", "w") as f:
        for i , key in enumerate(list(sorted(data.keys()))):
            if i>0:
                f.write("---\n")
            tp, date = key
            f.write("tipo: %s\nfecha: %s\narchivos:\n" % (tp, date.replace(".", "-")))
            for geo, url, out in sorted(data[key]):
                f.write("- %s\n- %s\n" % (url, out))


def save(file, data):
    with open(file, "w") as f:
        yaml.safe_dump_all(data, f, default_flow_style=False,
                           allow_unicode=True, default_style=None)

if args.analiza:
    for i in load("index.yml"):
        if i.tipo == "congreso":
            for archivo in i.archivos:
                if not archivo.startswith("http"):
                    yml = archivo.rsplit(".", 1)[0]+".yml"
                    if not os.path.isfile(yml) or args.machaca:
                        circunscripciones = parse_xlsx(archivo)
                        save(yml, circunscripciones)
